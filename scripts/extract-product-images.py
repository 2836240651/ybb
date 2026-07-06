#!/usr/bin/env python3
"""�?产品表单.xlsx 各工作区提取产品图片�?
命名规则（以货号为准）：
- 按表头识别「货�?/ 产品货号」与「产品编�?/ 产品编号」列（各工作区列位可不同�?- 默认�?parent 货号；同一货号多张不同图时改用 variant 产品编码
- 无货号列的工作区（如 2026新品）从 sku-mappings JSON �?TZ 货号
"""

from __future__ import annotations

import argparse
import json
import re
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from form_parser import allocate_mapped_skus, build_mapping_key, clean_text, load_sku_mapping

DEFAULT_XLSX = Path(r"C:\Users\Administrator\Pictures\excel表单图\产品表单.xlsx")
DEFAULT_OUTPUT = Path(r"C:\Users\Administrator\Pictures\excel表单�?)
_FALLBACK_SCHEMA = Path(
    r"d:\开发\fishAgent\总包\skill\reverse-skill\omc-replica\ybb-site\scripts\form-schema.json"
)
_LOCAL_SCHEMA = Path(__file__).resolve().parent / "form-schema.json"
SCHEMA_PATH = _LOCAL_SCHEMA if _LOCAL_SCHEMA.exists() else _FALLBACK_SCHEMA
SKU_MAPPINGS_DIR = Path(__file__).resolve().parents[1] / "deploy" / "product-import" / "sku-mappings"

DISPIMG_RE = re.compile(r'DISPIMG\s*\(\s*"([^"]+)"', re.IGNORECASE)
INVALID_FS_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
SKU_LIKE_RE = re.compile(r"^TZ-.+$", re.IGNORECASE)


def parse_dispimg(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    match = DISPIMG_RE.search(text)
    return match.group(1) if match else None


def should_skip_header_row(row: tuple[Any, ...], schema: dict[str, Any]) -> bool:
    first = clean_text(row[0] if row else "")
    aliases = schema.get("headerAliases", {}).get("skipFirstCell", [])
    return first in aliases or first.lower() == "name"


def is_parent_sku_header(value: str, schema: dict[str, Any]) -> bool:
    aliases = schema.get("headerAliases", {}).get("parentSku", [])
    return value in aliases


def is_variant_sku_header(value: str, schema: dict[str, Any]) -> bool:
    aliases = schema.get("headerAliases", {}).get("variantSku", [])
    return value in aliases


def cell_at(row: tuple[Any, ...], col_idx: int | None) -> Any:
    if col_idx is None or col_idx < 0 or col_idx >= len(row):
        return None
    return row[col_idx]


def resolve_sheet_columns(header_row: tuple[Any, ...], schema: dict[str, Any]) -> dict[str, int | None]:
    headers = [clean_text(cell) for cell in header_row]
    parent_aliases = set(schema.get("headerAliases", {}).get("parentSku", []))
    variant_aliases = set(schema.get("headerAliases", {}).get("variantSku", []))

    cols: dict[str, int | None] = {
        "name": 0,
        "image": 1,
        "spec": 2,
        "parent_sku": None,
        "variant_sku": None,
    }

    for idx, header in enumerate(headers):
        if not header:
            continue
        if header in parent_aliases:
            cols["parent_sku"] = idx
        elif header in variant_aliases:
            cols["variant_sku"] = idx
        elif header in {"图片", "Image", "image"}:
            cols["image"] = idx
        elif header in {"规格", "Spec", "spec"}:
            cols["spec"] = idx
        elif header in {"名称", "Name", "name"}:
            cols["name"] = idx

    return cols


def sanitize_filename(name: str, max_len: int = 120) -> str:
    text = INVALID_FS_CHARS.sub("_", name).strip(" .")
    text = text.rstrip(".")
    if not text:
        text = "unnamed"
    if len(text) > max_len:
        text = text[:max_len].rstrip(" .")
    return text


def is_valid_sku(value: str) -> bool:
    text = clean_text(value)
    return bool(text) and bool(SKU_LIKE_RE.match(text))


def unique_path(directory: Path, stem: str, suffix: str, used: set[str]) -> Path:
    candidate = f"{stem}{suffix}"
    if candidate not in used and not (directory / candidate).exists():
        used.add(candidate)
        return directory / candidate

    index = 2
    while True:
        alt = f"{stem}_{index}{suffix}"
        if alt not in used and not (directory / alt).exists():
            used.add(alt)
            return directory / alt
        index += 1


def build_dispimg_media_map(xlsx_path: Path) -> dict[str, str]:
    with zipfile.ZipFile(xlsx_path) as archive:
        try:
            cell_xml = archive.read("xl/cellimages.xml").decode("utf-8", errors="ignore")
            rels_xml = archive.read("xl/_rels/cellimages.xml.rels").decode("utf-8", errors="ignore")
        except KeyError as exc:
            raise SystemExit(
                "未找�?xl/cellimages.xml，该表可能不�?WPS DISPIMG 图片格式�?
            ) from exc

    rid_to_media = dict(
        re.findall(r'Id="(rId\d+)"[^>]*Target="media/([^"]+)"', rels_xml)
    )
    dispimg_to_media: dict[str, str] = {}
    for match in re.finditer(
        r'name="(ID_[A-F0-9]+)"[\s\S]*?r:embed="(rId\d+)"',
        cell_xml,
        flags=re.IGNORECASE,
    ):
        media = rid_to_media.get(match.group(2))
        if media:
            dispimg_to_media[match.group(1)] = media
    return dispimg_to_media


def lookup_mapped_skus(
    *,
    sheet_name: str,
    product_type: str,
    spec: str,
    color: str,
    mapping: dict[str, Any],
    include_color_in_key: bool,
) -> tuple[str, str]:
    key = build_mapping_key(
        sheet_name,
        product_type,
        spec,
        color if include_color_in_key else "",
    )
    entry = mapping.get("entries", {}).get(key)
    if entry:
        return clean_text(entry.get("parentSku")), clean_text(entry.get("variantSku"))

    if not spec:
        parent_sku = mapping.get("parentByType", {}).get(product_type)
        if parent_sku:
            parent_sku = clean_text(parent_sku)
            return parent_sku, parent_sku

    if product_type and spec:
        parent_sku, variant_sku, _ = allocate_mapped_skus(
            mapping,
            key=key,
            product_type=product_type,
            spec=spec,
        )
        return parent_sku, variant_sku

    if product_type:
        fallback_key = build_mapping_key(sheet_name, product_type, "Default", "")
        parent_sku, variant_sku, _ = allocate_mapped_skus(
            mapping,
            key=fallback_key,
            product_type=product_type,
            spec="Default",
        )
        return parent_sku, variant_sku

    return "", ""


def resolve_image_stem(
    *,
    parent_sku: str,
    variant_sku: str,
    parent_sku_seen: dict[str, str],
    dispimg_id: str,
) -> str | None:
    parent = clean_text(parent_sku)
    variant = clean_text(variant_sku)

    if is_valid_sku(parent):
        prior = parent_sku_seen.get(parent)
        if prior is None or prior == dispimg_id:
            parent_sku_seen[parent] = dispimg_id
            return sanitize_filename(parent)
        if is_valid_sku(variant):
            return sanitize_filename(variant)

    if is_valid_sku(variant):
        return sanitize_filename(variant)

    return None


def iter_sheet_image_rows(
    ws_formula,
    ws_values,
    *,
    sheet_name: str,
    sheet_cfg: dict[str, Any],
    schema: dict[str, Any],
    sku_mapping: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current_type = ""

    formula_rows = list(ws_formula.iter_rows(values_only=True))
    if not formula_rows:
        return rows

    header_row = formula_rows[0]
    cols = resolve_sheet_columns(header_row, schema)
    include_color_in_key = sheet_name == "2026新品"

    for row_index, row in enumerate(formula_rows, start=1):
        if row_index == 1 or should_skip_header_row(row, schema):
            continue

        value_rows = list(
            ws_values.iter_rows(min_row=row_index, max_row=row_index, values_only=True)
        )
        value_row = value_rows[0] if value_rows else ()

        name = clean_text(cell_at(row, cols["name"]))
        if name and not name.startswith("="):
            current_type = name

        spec = clean_text(cell_at(row, cols["spec"]))
        parent_sku = clean_text(cell_at(value_row, cols["parent_sku"]))
        variant_sku = clean_text(cell_at(value_row, cols["variant_sku"]))
        color = ""
        if include_color_in_key and cols["parent_sku"] is None:
            color = clean_text(cell_at(value_row, 3))
            if color.startswith("TZ-"):
                color = ""

        if parent_sku and is_parent_sku_header(parent_sku, schema):
            parent_sku = ""
        if variant_sku and is_variant_sku_header(variant_sku, schema):
            variant_sku = ""

        if not parent_sku and sku_mapping is not None:
            mapped_parent, mapped_variant = lookup_mapped_skus(
                sheet_name=sheet_name,
                product_type=current_type,
                spec=spec,
                color=color,
                mapping=sku_mapping,
                include_color_in_key=include_color_in_key,
            )
            parent_sku = parent_sku or mapped_parent
            variant_sku = variant_sku or mapped_variant

        dispimg_id = parse_dispimg(cell_at(row, cols["image"]))
        if not dispimg_id:
            continue
        if not current_type and not spec:
            continue

        rows.append(
            {
                "sheet": sheet_name,
                "row": row_index,
                "product_type": current_type,
                "spec": spec,
                "parent_sku": parent_sku,
                "variant_sku": variant_sku,
                "dispimg_id": dispimg_id,
            }
        )

    return rows


def extract_all_product_images(
    excel_path: str | Path,
    output_dir: str | Path = DEFAULT_OUTPUT,
    *,
    schema_path: Path = SCHEMA_PATH,
    per_sheet_subdir: bool = False,
    sku_mappings_dir: Path = SKU_MAPPINGS_DIR,
    only_sheets: list[str] | None = None,
) -> dict[str, int]:
    xlsx_path = Path(excel_path)
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel 文件不存�? {xlsx_path}")

    schema = json.loads(schema_path.read_text(encoding="utf-8-sig"))
    sheet_map: dict[str, Any] = schema.get("sheetMap", {})
    skip_sheets = set(schema.get("skipSheets", []))

    dispimg_to_media = build_dispimg_media_map(xlsx_path)
    wb_formula = load_workbook(xlsx_path, data_only=False)
    wb_values = load_workbook(xlsx_path, data_only=True)

    stats = {
        "total": 0,
        "ok": 0,
        "missing_media": 0,
        "missing_sheet": 0,
        "skipped_dup": 0,
        "skipped_no_sku": 0,
    }
    used_names: set[str] = set()
    parent_sku_seen: dict[str, str] = {}
    dispimg_saved: set[str] = set()

    with zipfile.ZipFile(xlsx_path) as archive:
        for sheet_name, sheet_cfg in sheet_map.items():
            if only_sheets is not None and sheet_name not in only_sheets:
                continue
            if sheet_name in skip_sheets:
                continue
            if sheet_name not in wb_formula.sheetnames:
                stats["missing_sheet"] += 1
                print(f"[跳过] 工作表不存在: {sheet_name}")
                continue

            ws_formula = wb_formula[sheet_name]
            ws_values = wb_values[sheet_name]
            target_dir = out_dir / sheet_name if per_sheet_subdir else out_dir
            target_dir.mkdir(parents=True, exist_ok=True)

            sku_mapping = None
            mapping_file = sheet_cfg.get("skuMappingFile")
            if mapping_file:
                prefix = str(sheet_cfg.get("skuPrefix") or "TZ-XX")
                sku_mapping = load_sku_mapping(sku_mappings_dir / mapping_file, prefix)

            image_rows = iter_sheet_image_rows(
                ws_formula,
                ws_values,
                sheet_name=sheet_name,
                sheet_cfg=sheet_cfg,
                schema=schema,
                sku_mapping=sku_mapping,
            )

            for item in image_rows:
                stats["total"] += 1
                if item["dispimg_id"] in dispimg_saved:
                    stats["skipped_dup"] += 1
                    continue

                stem = resolve_image_stem(
                    parent_sku=item["parent_sku"],
                    variant_sku=item["variant_sku"],
                    parent_sku_seen=parent_sku_seen,
                    dispimg_id=item["dispimg_id"],
                )
                if not stem:
                    stats["skipped_no_sku"] += 1
                    print(
                        f"[无货号] {item['sheet']} 行{item['row']} "
                        f"type={item['product_type']} spec={item['spec']}"
                    )
                    continue

                media_name = dispimg_to_media.get(item["dispimg_id"])
                if not media_name:
                    stats["missing_media"] += 1
                    print(
                        f"[缺失] {item['sheet']} 行{item['row']} "
                        f"{stem} -> {item['dispimg_id']}"
                    )
                    continue

                media_path = f"xl/media/{media_name}"
                try:
                    data = archive.read(media_path)
                except KeyError:
                    stats["missing_media"] += 1
                    print(
                        f"[缺失] {item['sheet']} 行{item['row']} "
                        f"{stem} -> {media_path}"
                    )
                    continue

                suffix = Path(media_name).suffix.lower() or ".png"
                dest = unique_path(target_dir, stem, suffix, used_names)
                dest.write_bytes(data)
                dispimg_saved.add(item["dispimg_id"])
                stats["ok"] += 1
                label = item["parent_sku"] or item["variant_sku"] or stem
                print(f"提取: {dest.name}  ({item['sheet']} / {label})")

    return stats


def main() -> int:
    parser = argparse.ArgumentParser(description="提取产品表单图片（按表头识别货号列）�?)
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--schema", default=str(SCHEMA_PATH))
    parser.add_argument("--sku-mappings-dir", default=str(SKU_MAPPINGS_DIR))
    parser.add_argument("--only-sheets", nargs="+", help="Only extract these worksheet names")
    parser.add_argument("--per-sheet-subdir", action="store_true")
    parser.add_argument("--clean-output", action="store_true", help="提取前清空输出目录内已有图片")
    args = parser.parse_args()

    out_dir = Path(args.output)
    if args.clean_output and out_dir.exists():
        removed = 0
        for path in out_dir.iterdir():
            if path.is_file() and path.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp", ".gif"}:
                path.unlink()
                removed += 1
        print(f"已清空旧图片: {removed} �?)

    stats = extract_all_product_images(
        args.xlsx,
        args.output,
        schema_path=Path(args.schema),
        per_sheet_subdir=args.per_sheet_subdir,
        sku_mappings_dir=Path(args.sku_mappings_dir),
        only_sheets=args.only_sheets,
    )

    print(
        "提取完成: "
        f"�?{stats['total']} �? 成功 {stats['ok']}, "
        f"重复跳过 {stats['skipped_dup']}, 无货�?{stats['skipped_no_sku']}, "
        f"缺失 {stats['missing_media']}, 缺表 {stats['missing_sheet']}"
    )
    return 0 if stats["missing_media"] == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
