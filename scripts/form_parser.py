#!/usr/bin/env python3
"""Parse 产品表单.xlsx into sellable WooCommerce simple-product rows."""

from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPT_DIR = Path(__file__).resolve().parent
SCHEMA_PATH = SCRIPT_DIR / "form-schema.json"
DEFAULT_XLSX = Path.home() / "Desktop" / "产品表单.xlsx"
DEFAULT_OUT = ROOT / "deploy" / "product-import"
SKU_MAPPINGS_DIR = DEFAULT_OUT / "sku-mappings"

DISPIMG_RE = re.compile(r'=DISPIMG\s*\(\s*"([^"]+)"', re.IGNORECASE)
SKU_SAFE_RE = re.compile(r"[^\w\-+#./]+", re.UNICODE)

CSV_HEADER = [
    "Type",
    "SKU",
    "Name",
    "Published",
    "Visibility in catalog",
    "Tax status",
    "In stock?",
    "Regular price",
    "Categories",
    "Grouped",
    "Attribute 1 name",
    "Attribute 1 value(s)",
    "Attribute 1 visible",
    "Attribute 1 global",
    "Meta: _parent_sku",
    "Meta: _sheet",
    "Meta: _mapping_key",
    "Meta: _ybb_title_zh",
]


def load_schema(path: Path | None = None) -> dict[str, Any]:
    schema_path = path or SCHEMA_PATH
    return json.loads(schema_path.read_text(encoding="utf-8-sig"))


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip()
    return re.sub(r"\s+", " ", text)


def price_to_string(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        value = value.strip().replace("$", "").replace("�?, "").replace("¥", "")
        if not value:
            return None
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return None
    formatted = f"{amount:.2f}".rstrip("0").rstrip(".")
    return formatted or "0"


def parse_dispimg(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    match = DISPIMG_RE.search(text)
    return match.group(1) if match else None


def slugify_type(parent_handle: str, type_name: str, index: int) -> str:
    base = re.sub(r"\s+", "-", type_name.strip().lower())
    base = re.sub(r"[^\w\-]+", "", base, flags=re.UNICODE)
    base = base.strip("-")[:40]
    if not base:
        base = f"type-{index}"
    return f"{parent_handle}--{base}"


def sanitize_spec_for_sku(spec: str) -> str:
    from sku_normalize import sanitize_spec_for_sku as _canonical_sanitize

    return _canonical_sanitize(spec)


def sku_handle(sku: str) -> str:
    text = (sku or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def resolve_price_col(header_row: tuple[Any, ...], price_columns: list[dict[str, Any]] | None) -> int | None:
    if not price_columns:
        return None
    headers = [clean_text(cell) for cell in header_row]
    for config in price_columns:
        for alias in config.get("headers", []):
            alias = clean_text(alias)
            for idx, header in enumerate(headers):
                if header == alias or alias in header:
                    return idx
        if "col" in config:
            return int(config["col"])
    return None


def should_skip_header_row(row: tuple[Any, ...], schema: dict[str, Any]) -> bool:
    first = clean_text(row[0] if row else "")
    aliases = schema.get("headerAliases", {}).get("skipFirstCell", [])
    return first in aliases or first.lower() == "name"


def is_parent_sku_header(value: str, schema: dict[str, Any]) -> bool:
    aliases = schema.get("headerAliases", {}).get("parentSku", [])
    return value in aliases


def build_mapping_key(sheet: str, product_type: str, spec: str, color: str = "") -> str:
    return "|".join([sheet, product_type, spec, color])


def default_sku_mapping(prefix: str) -> dict[str, Any]:
    return {
        "version": 1,
        "prefix": prefix,
        "nextSequence": 1,
        "parentByType": {},
        "entries": {},
    }


def load_sku_mapping(path: Path, prefix: str) -> dict[str, Any]:
    if path.exists():
        data = json.loads(path.read_text(encoding="utf-8"))
        data.setdefault("version", 1)
        data.setdefault("prefix", prefix)
        data.setdefault("nextSequence", 1)
        data.setdefault("parentByType", {})
        data.setdefault("entries", {})
        return data
    return default_sku_mapping(prefix)


def allocate_mapped_skus(
    mapping: dict[str, Any],
    *,
    key: str,
    product_type: str,
    spec: str,
) -> tuple[str, str, bool]:
    entries: dict[str, dict[str, str]] = mapping.setdefault("entries", {})
    if key in entries:
        entry = entries[key]
        return entry["parentSku"], entry["variantSku"], False

    prefix = mapping["prefix"]
    parent_by_type: dict[str, str] = mapping.setdefault("parentByType", {})
    if product_type not in parent_by_type:
        seq = int(mapping.get("nextSequence", 1))
        parent_by_type[product_type] = f"{prefix}-{seq:03d}"
        mapping["nextSequence"] = seq + 1

    parent_sku = parent_by_type[product_type]
    variant_sku = f"{parent_sku}-{sanitize_spec_for_sku(spec)}"
    entries[key] = {"parentSku": parent_sku, "variantSku": variant_sku, "spec": spec}
    return parent_sku, variant_sku, True


@dataclass
class SellableRow:
    sheet: str
    sheet_handle: str
    sheet_title_en: str
    sheet_title_ja: str
    row_index: int
    product_type_name: str
    spec: str
    parent_sku: str
    variant_sku: str
    price: str
    image_ref: str | None = None
    color: str = ""
    mapping_key: str = ""
    category_slugs: list[str] = field(default_factory=list)
    needs_review: bool = False
    price_is_default: bool = False
    published: str = "1"

    @property
    def name_en(self) -> str:
        """Woo / EN storefront title (ASCII-friendly, no Chinese type name)."""
        parts = [self.sheet_title_en]
        if self.spec:
            parts.append(self.spec)
        if self.color:
            parts.append(self.color)
        return " - ".join(part for part in parts if part)

    @property
    def name_ja(self) -> str:
        parts = [self.sheet_title_ja]
        if self.spec:
            parts.append(self.spec)
        if self.color:
            parts.append(self.color)
        return " - ".join(part for part in parts if part)

    @property
    def name_zh(self) -> str:
        """ZH display title from form."""
        parts = [self.product_type_name or self.sheet]
        if self.spec:
            parts.append(self.spec)
        if self.color:
            parts.append(self.color)
        return " - ".join(part for part in parts if part)

    @property
    def name(self) -> str:
        return self.name_en

    @property
    def categories(self) -> str:
        return " > ".join(self.category_slugs)

    def parent_name_en(self) -> str:
        parts = [self.sheet_title_en]
        if self.product_type_name:
            parts.append(self.product_type_name)
        return " - ".join(part for part in parts if part)

    def parent_name_ja(self) -> str:
        parts = [self.sheet_title_ja]
        if self.product_type_name:
            parts.append(self.product_type_name)
        return " - ".join(part for part in parts if part)

    def parent_name_zh(self) -> str:
        return self.product_type_name or self.sheet

    def to_csv_row(self) -> list[str]:
        return [
            "simple",
            self.variant_sku,
            self.name,
            self.published,
            "visible",
            "taxable",
            "1",
            self.price,
            self.categories,
            self.parent_sku,
            "Spec",
            self.spec or "Default",
            "1",
            "0",
            self.parent_sku,
            self.sheet,
            self.mapping_key,
            self.name_zh,
        ]


@dataclass
class CatalogVariation:
    sku: str
    spec: str
    price: str
    titleZh: str = ""
    titleJa: str = ""
    imageRef: str | None = None


@dataclass
class CatalogProduct:
    parentSku: str
    name: str
    nameZh: str
    nameJa: str
    type: str
    categorySlugs: list[str]
    sheet: str
    sheetHandle: str
    variations: list[CatalogVariation]
    regularPrice: str | None = None
    variationSku: str | None = None
    spec: str | None = None


def group_rows_by_parent(rows: list[SellableRow]) -> list[CatalogProduct]:
    grouped: dict[str, list[SellableRow]] = {}
    for row in rows:
        grouped.setdefault(row.parent_sku, []).append(row)

    products: list[CatalogProduct] = []
    for parent_sku in sorted(grouped):
        items = sorted(grouped[parent_sku], key=lambda item: item.row_index)
        first = items[0]
        variations = [
            CatalogVariation(
                sku=item.variant_sku,
                spec=item.spec or "Default",
                price=item.price,
                titleZh=item.name_zh,
                titleJa=item.name_ja,
                imageRef=item.image_ref,
            )
            for item in items
        ]
        base = {
            "parentSku": parent_sku,
            "name": first.parent_name_en(),
            "nameZh": first.parent_name_zh(),
            "nameJa": first.parent_name_ja(),
            "categorySlugs": list(first.category_slugs),
            "sheet": first.sheet,
            "sheetHandle": first.sheet_handle,
            "variations": variations,
        }
        if len(items) >= 2:
            products.append(CatalogProduct(type="variable", **base))
        else:
            only = variations[0]
            products.append(
                CatalogProduct(
                    type="simple",
                    regularPrice=only.price,
                    variationSku=only.sku,
                    spec=only.spec,
                    **base,
                )
            )
    return products


def catalog_stats(products: list[CatalogProduct]) -> dict[str, int]:
    variation_count = sum(
        len(product.variations) for product in products if product.type == "variable"
    )
    return {
        "productCount": len(products),
        "variationCount": variation_count,
        "simpleCount": sum(1 for product in products if product.type == "simple"),
        "variableCount": sum(1 for product in products if product.type == "variable"),
    }


def catalog_product_to_dict(product: CatalogProduct) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "parentSku": product.parentSku,
        "name": product.name,
        "nameZh": product.nameZh,
        "nameJa": product.nameJa,
        "type": product.type,
        "categorySlugs": product.categorySlugs,
        "sheet": product.sheet,
        "sheetHandle": product.sheetHandle,
        "variations": [asdict(variation) for variation in product.variations],
    }
    if product.type == "simple":
        payload["regularPrice"] = product.regularPrice
        payload["variationSku"] = product.variationSku
        payload["spec"] = product.spec
    return payload


def build_variant_redirects(products: list[CatalogProduct]) -> dict[str, str]:
    redirects: dict[str, str] = {}
    for product in products:
        parent_handle = sku_handle(product.parentSku)
        if product.type != "variable":
            continue
        for variation in product.variations:
            variant_handle = sku_handle(variation.sku)
            if variant_handle and variant_handle != parent_handle:
                redirects[variant_handle] = parent_handle
    return redirects


def build_category_terms(
    wb,
    schema: dict[str, Any],
    rows: list[SellableRow],
) -> dict[str, Any]:
    sheet_map: dict[str, dict[str, Any]] = schema["sheetMap"]
    skip_sheets = set(schema.get("skipSheets", []))
    main_sheet_count = int(schema.get("mainSheetCount", 8))

    terms: list[dict[str, Any]] = [
        {"slug": "other", "nameEn": "Other", "nameZh": "其他", "parent": None},
    ]
    sku_to_slugs: dict[str, list[str]] = {}

    sheet_order = [name for name in wb.sheetnames if name not in skip_sheets and name in sheet_map]
    main_sheets = sheet_order[:main_sheet_count]
    other_sheets = sheet_order[main_sheet_count:]

    def add_sheet_terms(sheet_name: str, parent_override: str | None = None) -> str:
        sheet_cfg = sheet_map[sheet_name]
        handle = sheet_cfg["handle"]
        title_en = sheet_cfg["titleEn"]
        ws = wb[sheet_name]

        if parent_override is None:
            terms.append(
                {
                    "slug": handle,
                    "nameEn": title_en,
                    "nameZh": sheet_name,
                    "parent": None,
                }
            )
            parent_for_types = handle
        else:
            terms.append(
                {
                    "slug": handle,
                    "nameEn": title_en,
                    "nameZh": sheet_name,
                    "parent": parent_override,
                }
            )
            parent_for_types = handle

        seen: set[str] = set()
        current_type: str | None = None
        current_type_slug: str | None = None
        type_index = 0

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or should_skip_header_row(row, schema):
                continue
            col_a = row[0] if row else None
            if col_a is not None and clean_text(col_a) and not clean_text(col_a).startswith("="):
                current_type = clean_text(col_a)
                if current_type not in seen:
                    seen.add(current_type)
                    type_index += 1
                    current_type_slug = slugify_type(parent_for_types, current_type, type_index)
                    terms.append(
                        {
                            "slug": current_type_slug,
                            "nameEn": current_type,
                            "nameZh": current_type,
                            "parent": parent_for_types,
                        }
                    )

            if sheet_cfg.get("layout") == "standard":
                parent_sku = clean_text(row[3] if row and len(row) > 3 else None)
                if not parent_sku or is_parent_sku_header(parent_sku, schema):
                    continue
                if not current_type or not current_type_slug:
                    continue
                path = (
                    [parent_override, parent_for_types, current_type_slug]
                    if parent_override
                    else [parent_for_types, current_type_slug]
                )
                sku_to_slugs[parent_sku] = path

        return handle

    for sheet_name in main_sheets:
        add_sheet_terms(sheet_name)

    for sheet_name in other_sheets:
        add_sheet_terms(sheet_name, parent_override="other")

    for row in rows:
        if row.parent_sku and row.category_slugs:
            sku_to_slugs.setdefault(row.parent_sku, row.category_slugs)

    return {
        "taxonomy": "product_cat",
        "preserve": ["product_brand", "product_tag", "pa_*"],
        "removeLegacySlugs": schema.get("legacyProductCatSlugs", []),
        "skuPrefixFallback": {
            "TZ-XP": ["2026-new-products"],
            "TZ-HK": ["carp-hooks"],
        },
        "terms": terms,
        "skuToCategorySlugs": sku_to_slugs,
    }


def category_slugs_for_row(
    *,
    sheet_handle: str,
    product_type_slug: str | None,
    parent_override: str | None,
) -> list[str]:
    if not product_type_slug:
        if parent_override:
            return [parent_override, sheet_handle]
        return [sheet_handle]
    if parent_override:
        return [parent_override, sheet_handle, product_type_slug]
    return [sheet_handle, product_type_slug]


def parse_standard_sheet(
    ws,
    *,
    sheet_name: str,
    sheet_cfg: dict[str, Any],
    schema: dict[str, Any],
    parent_override: str | None,
    type_slug_by_name: dict[str, str],
) -> tuple[list[SellableRow], list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[SellableRow] = []
    blocked_rows: list[dict[str, Any]] = []
    image_refs: list[dict[str, Any]] = []

    handle = sheet_cfg["handle"]
    title_en = sheet_cfg["titleEn"]
    title_ja = sheet_cfg.get("titleJa", title_en)
    default_price = schema.get("defaultPrice", "1.99")
    price_columns = sheet_cfg.get("priceColumns")

    header_row: tuple[Any, ...] | None = None
    price_col: int | None = None
    current_type = ""
    current_type_slug: str | None = None

    for row_index, row in enumerate(ws.iter_rows(values_only=True)):
        if row_index == 0:
            header_row = row
            price_col = resolve_price_col(row, price_columns)
            continue
        if should_skip_header_row(row, schema):
            continue

        name = clean_text(row[0] if len(row) > 0 else None)
        if name and not name.startswith("="):
            current_type = name
            current_type_slug = type_slug_by_name.get(current_type)

        spec = clean_text(row[2] if len(row) > 2 else None)
        parent_sku = clean_text(row[3] if len(row) > 3 else None)
        variant_sku = clean_text(row[4] if len(row) > 4 else None)

        if not spec and not parent_sku:
            continue
        if not parent_sku or is_parent_sku_header(parent_sku, schema):
            blocked_rows.append(
                {
                    "sheet": sheet_name,
                    "row": row_index + 1,
                    "reason": "missing_parent_sku",
                    "spec": spec,
                    "productType": current_type,
                }
            )
            continue

        if not variant_sku:
            variant_sku = f"{parent_sku}-{sanitize_spec_for_sku(spec)}" if spec else parent_sku

        image_ref = parse_dispimg(row[1] if len(row) > 1 else None)
        if image_ref:
            image_refs.append(
                {
                    "sheet": sheet_name,
                    "row": row_index + 1,
                    "dispimgId": image_ref,
                    "parentSku": parent_sku,
                    "variantSku": variant_sku,
                }
            )

        price_value = row[price_col] if price_col is not None and len(row) > price_col else None
        price = price_to_string(price_value) or default_price
        default_applied = price_to_string(price_value) is None

        rows.append(
            SellableRow(
                sheet=sheet_name,
                sheet_handle=handle,
                sheet_title_en=title_en,
                sheet_title_ja=title_ja,
                row_index=row_index + 1,
                product_type_name=current_type,
                spec=spec,
                parent_sku=parent_sku,
                variant_sku=variant_sku,
                price=price,
                image_ref=image_ref,
                category_slugs=category_slugs_for_row(
                    sheet_handle=handle,
                    product_type_slug=current_type_slug,
                    parent_override=parent_override,
                ),
                price_is_default=default_applied,
            )
        )

    return rows, blocked_rows, image_refs


def parse_mapped_sheet(
    ws,
    *,
    sheet_name: str,
    sheet_cfg: dict[str, Any],
    schema: dict[str, Any],
    mapping: dict[str, Any],
    parent_override: str | None,
    type_slug_by_name: dict[str, str],
    include_color_in_key: bool,
) -> tuple[list[SellableRow], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    rows: list[SellableRow] = []
    blocked_rows: list[dict[str, Any]] = []
    image_refs: list[dict[str, Any]] = []

    handle = sheet_cfg["handle"]
    title_en = sheet_cfg["titleEn"]
    title_ja = sheet_cfg.get("titleJa", title_en)
    default_price = schema.get("defaultPrice", "1.99")
    price_columns = sheet_cfg.get("priceColumns")

    price_col: int | None = None
    current_type = ""
    current_type_slug: str | None = None

    for row_index, row in enumerate(ws.iter_rows(values_only=True)):
        if row_index == 0:
            price_col = resolve_price_col(row, price_columns)
            continue
        if should_skip_header_row(row, schema):
            continue

        name = clean_text(row[0] if len(row) > 0 else None)
        if name and not name.startswith("="):
            current_type = name
            current_type_slug = type_slug_by_name.get(current_type)

        spec = clean_text(row[2] if len(row) > 2 else None)
        color = clean_text(row[3] if len(row) > 3 else None)
        if color.startswith("TZ-"):
            color = ""

        if not spec:
            continue
        if not current_type:
            blocked_rows.append(
                {
                    "sheet": sheet_name,
                    "row": row_index + 1,
                    "reason": "missing_product_type",
                    "spec": spec,
                }
            )
            continue

        key = build_mapping_key(
            sheet_name,
            current_type,
            spec,
            color if include_color_in_key else "",
        )
        parent_sku, variant_sku, is_new = allocate_mapped_skus(
            mapping,
            key=key,
            product_type=current_type,
            spec=spec,
        )

        image_ref = parse_dispimg(row[1] if len(row) > 1 else None)
        if image_ref:
            image_refs.append(
                {
                    "sheet": sheet_name,
                    "row": row_index + 1,
                    "dispimgId": image_ref,
                    "parentSku": parent_sku,
                    "variantSku": variant_sku,
                }
            )

        price_value = row[price_col] if price_col is not None and len(row) > price_col else None
        parsed_price = price_to_string(price_value)
        price = parsed_price or default_price

        rows.append(
            SellableRow(
                sheet=sheet_name,
                sheet_handle=handle,
                sheet_title_en=title_en,
                sheet_title_ja=title_ja,
                row_index=row_index + 1,
                product_type_name=current_type,
                spec=spec,
                parent_sku=parent_sku,
                variant_sku=variant_sku,
                price=price,
                image_ref=image_ref,
                color=color,
                mapping_key=key,
                category_slugs=category_slugs_for_row(
                    sheet_handle=handle,
                    product_type_slug=current_type_slug,
                    parent_override=parent_override,
                ),
                needs_review=is_new,
                price_is_default=parsed_price is None,
            )
        )

    return rows, blocked_rows, image_refs, mapping


def type_slug_index_for_sheet(wb, sheet_name: str, sheet_cfg: dict[str, Any], schema: dict[str, Any]) -> dict[str, str]:
    ws = wb[sheet_name]
    handle = sheet_cfg["handle"]
    seen: set[str] = set()
    current_type = ""
    type_index = 0
    mapping: dict[str, str] = {}

    for row_index, row in enumerate(ws.iter_rows(values_only=True)):
        if row_index == 0 or should_skip_header_row(row, schema):
            continue
        name = clean_text(row[0] if row else None)
        if name and not name.startswith("="):
            current_type = name
            if current_type not in seen:
                seen.add(current_type)
                type_index += 1
                mapping[current_type] = slugify_type(handle, current_type, type_index)
    return mapping


def parse_workbook(
    xlsx_path: Path,
    *,
    schema: dict[str, Any] | None = None,
    sku_mappings_dir: Path | None = None,
) -> tuple[list[SellableRow], dict[str, Any]]:
    schema = schema or load_schema()
    sku_mappings_dir = sku_mappings_dir or SKU_MAPPINGS_DIR
    sheet_map: dict[str, dict[str, Any]] = schema["sheetMap"]
    skip_sheets = set(schema.get("skipSheets", []))
    main_sheet_count = int(schema.get("mainSheetCount", 8))
    default_price = schema.get("defaultPrice", "1.99")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    sheet_order = [name for name in wb.sheetnames if name not in skip_sheets and name in sheet_map]
    main_sheets = sheet_order[:main_sheet_count]
    other_sheets = sheet_order[main_sheet_count:]

    loaded_mappings: dict[str, dict[str, Any]] = {}
    all_rows: list[SellableRow] = []
    blocked_rows: list[dict[str, Any]] = []
    image_refs: list[dict[str, Any]] = []
    stats_by_sheet: dict[str, dict[str, int]] = {}

    def parse_one(sheet_name: str, parent_override: str | None) -> None:
        sheet_cfg = sheet_map[sheet_name]
        ws = wb[sheet_name]
        type_slugs = type_slug_index_for_sheet(wb, sheet_name, sheet_cfg, schema)
        layout = sheet_cfg.get("layout", "standard")

        if layout == "mapped":
            mapping_file = sheet_cfg["skuMappingFile"]
            mapping_path = sku_mappings_dir / mapping_file
            prefix = sheet_cfg["skuPrefix"]
            mapping = load_sku_mapping(mapping_path, prefix)
            include_color = sheet_name == "2026新品"
            sheet_rows, sheet_blocked, sheet_images, updated_mapping = parse_mapped_sheet(
                ws,
                sheet_name=sheet_name,
                sheet_cfg=sheet_cfg,
                schema=schema,
                mapping=mapping,
                parent_override=parent_override,
                type_slug_by_name=type_slugs,
                include_color_in_key=include_color,
            )
            loaded_mappings[mapping_file] = updated_mapping
        else:
            sheet_rows, sheet_blocked, sheet_images = parse_standard_sheet(
                ws,
                sheet_name=sheet_name,
                sheet_cfg=sheet_cfg,
                schema=schema,
                parent_override=parent_override,
                type_slug_by_name=type_slugs,
            )

        all_rows.extend(sheet_rows)
        blocked_rows.extend(sheet_blocked)
        image_refs.extend(sheet_images)
        stats_by_sheet[sheet_name] = {
            "sellableRows": len(sheet_rows),
            "blockedRows": len(sheet_blocked),
            "needsReview": sum(1 for item in sheet_rows if item.needs_review),
        }

    for sheet_name in main_sheets:
        parse_one(sheet_name, None)
    for sheet_name in other_sheets:
        parse_one(sheet_name, "other")

    category_terms = build_category_terms(wb, schema, all_rows)
    for row in all_rows:
        if row.parent_sku in category_terms["skuToCategorySlugs"]:
            row.category_slugs = category_terms["skuToCategorySlugs"][row.parent_sku]

    default_price_applied = sum(1 for row in all_rows if row.price_is_default)
    catalog_products = group_rows_by_parent(all_rows)
    catalog_counts = catalog_stats(catalog_products)

    manifest = {
        "source": str(xlsx_path),
        "sourceSha256": sha256_file(xlsx_path),
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "defaultPrice": default_price,
        "attribute": {"name": "Spec", "slug": "spec", "taxonomy": "pa_spec"},
        "categoryTerms": category_terms,
        "imageRefs": image_refs,
        "blockedRows": blocked_rows,
        "needsReview": [asdict(row) for row in all_rows if row.needs_review],
        "skuMappingFiles": loaded_mappings,
        "i18nBySku": {
            row.variant_sku: {"titleEn": row.name_en, "titleZh": row.name_zh, "titleJa": row.name_ja}
            for row in all_rows
        },
        "i18nByParentSku": {
            product.parentSku: {"titleEn": product.name, "titleZh": product.nameZh, "titleJa": product.nameJa}
            for product in catalog_products
        },
        "stats": {
            "sellableSkuCount": len(all_rows),
            "blockedRowCount": len(blocked_rows),
            "imageRefCount": len(image_refs),
            "defaultPriceAppliedCount": default_price_applied,
            "needsReviewCount": sum(1 for row in all_rows if row.needs_review),
            "sheetStats": stats_by_sheet,
            "termCount": len(category_terms["terms"]),
            "skuBaseCount": len(category_terms["skuToCategorySlugs"]),
            **catalog_counts,
        },
    }
    return all_rows, manifest


def write_csv(rows: list[SellableRow], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(CSV_HEADER)
        for row in rows:
            writer.writerow(row.to_csv_row())


def write_manifest(manifest: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_sku_mappings(manifest: dict[str, Any], directory: Path | None = None) -> None:
    directory = directory or SKU_MAPPINGS_DIR
    directory.mkdir(parents=True, exist_ok=True)
    for filename, payload in manifest.get("skuMappingFiles", {}).items():
        target = directory / filename
        target.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_product_i18n(manifest: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    merged = {
        **manifest.get("i18nByParentSku", {}),
        **manifest.get("i18nBySku", {}),
    }
    path.write_text(
        json.dumps(merged, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def write_wc_catalog(rows: list[SellableRow], path: Path) -> list[CatalogProduct]:
    products = group_rows_by_parent(rows)
    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "stats": catalog_stats(products),
        "products": [catalog_product_to_dict(product) for product in products],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return products


def write_variant_redirects(products: list[CatalogProduct], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(build_variant_redirects(products), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
