#!/usr/bin/env python3
"""Generate the single-sheet Excel template for carp-ybb product listing."""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "deploy" / "product-import" / "templates" / "carp-ybb_product-listing-template-single-sheet.xlsx"

CATEGORIES = [
    "2026 新品 (2026 New Products)",
    "铅坠 (Sinkers)",
    "饵笼 (Bait Cages)",
    "钓组 (Rigs)",
    "铅坠钓组 (Sinker Rigs)",
    "饵笼钓组 (Bait Cage Rigs)",
    "欧鲤钩 (Carp Hooks)",
    "欧鲤套装 (Euro Carp Kits)",
    "其他 (Other)",
]

STATUS_OPTIONS = ["草稿 (draft)", "发布 (publish)"]
REVIEW_STATUS_OPTIONS = ["已批准 (approved)", "待审核 (hold)"]

LABELS = {
    "parent_sku": "父SKU (parent_sku)",
    "handle": "前台路径slug (handle)",
    "title_en": "英文标题 (title_en)",
    "title_zh": "中文标题 (title_zh)",
    "title_ja": "日文标题 (title_ja)",
    "description_en_html": "英文描述HTML (description_en_html)",
    "description_zh_html": "中文描述HTML (description_zh_html)",
    "description_ja_html": "日文描述HTML (description_ja_html)",
    "category": "类目 (category)",
    "subcategory": "子类目 (subcategory)",
    "status": "上架状态 (status)",
    "tag_csv": "标签，逗号分隔 (tag_csv)",
    "attribute_name": "规格属性名 (attribute_name)",
    "main_image_url": "主图URL (main_image_url)",
    "gallery_image_url_1": "图库URL 1 (gallery_image_url_1)",
    "gallery_image_url_2": "图库URL 2 (gallery_image_url_2)",
    "gallery_image_url_3": "图库URL 3 (gallery_image_url_3)",
    "gallery_image_url_4": "图库URL 4 (gallery_image_url_4)",
    "gallery_image_url_5": "图库URL 5 (gallery_image_url_5)",
    "gallery_image_url_6": "图库URL 6 (gallery_image_url_6)",
    "local_image_path": "本地图片路径 (local_image_path)",
    "image_alt_en": "图片英文ALT (image_alt_en)",
    "image_alt_zh": "图片中文ALT (image_alt_zh)",
    "image_alt_ja": "图片日文ALT (image_alt_ja)",
    "image_ready_0_1": "图片已齐备 0/1 (image_ready_0_1)",
    "hide_description_0_1": "隐藏描述 0/1 (hide_description_0_1)",
    "hide_additional_info_0_1": "隐藏附加信息 0/1 (hide_additional_info_0_1)",
    "front_hidden_0_1": "前台隐藏 0/1 (front_hidden_0_1)",
    "gallery_enabled_0_1": "启用图库 0/1 (gallery_enabled_0_1)",
    "gallery_default_index": "默认图库序号 (gallery_default_index)",
    "gallery_override_enabled_0_1": "启用图库覆盖 0/1 (gallery_override_enabled_0_1)",
    "gallery_hide_indexes": "隐藏图库序号 (gallery_hide_indexes)",
    "slogan_en": "英文购买区文案 (slogan_en)",
    "slogan_zh": "中文购买区文案 (slogan_zh)",
    "slogan_ja": "日文购买区文案 (slogan_ja)",
    "hide_slogan_0_1": "隐藏购买区文案 0/1 (hide_slogan_0_1)",
    "review_author": "评价人 (review_author)",
    "review_email": "评价邮箱 (review_email)",
    "review_rating_1_5": "评分1-5 (review_rating_1_5)",
    "review_content": "评价内容 (review_content)",
    "review_date_yyyy_mm_dd": "评价日期YYYY-MM-DD (review_date_yyyy_mm_dd)",
    "review_status": "评价状态 (review_status)",
    "ops_note": "运营备注 (ops_note)",
}

for index in range(1, 7):
    LABELS.update(
        {
            f"variant{index}_sku": f"变体{index} SKU (variant{index}_sku)",
            f"variant{index}_parent_sku": f"变体{index} 父SKU (variant{index}_parent_sku)",
            f"variant{index}_spec_value": f"变体{index} 规格值 (variant{index}_spec_value)",
            f"variant{index}_price_usd": f"变体{index} 美元价格 (variant{index}_price_usd)",
            f"variant{index}_in_stock_0_1": f"变体{index} 有货0/1 (variant{index}_in_stock_0_1)",
            f"variant{index}_image_url": f"变体{index} 图片URL (variant{index}_image_url)",
        }
    )


def _list_validation(options: list[str], allow_blank: bool = True) -> DataValidation:
    return DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=allow_blank)


def _apply_list(ws, cell_range: str, options: list[str], allow_blank: bool = True) -> None:
    dv = _list_validation(options, allow_blank=allow_blank)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def _field_range(columns: list[str], field: str, start_row: int = 3, end_row: int = 1000) -> str:
    col = get_column_letter(columns.index(field) + 1)
    return f"{col}{start_row}:{col}{end_row}"


def _style_group_row(ws, groups: list[tuple[int, int, str]]) -> None:
    fills = ["334155", "155E75", "166534", "7C2D12", "581C87"]
    for idx, (start_col, end_col, label) in enumerate(groups):
        fill = PatternFill("solid", fgColor=fills[idx % len(fills)])
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=start_col, value=label)


def _style_header_row(ws) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(vertical="center", wrap_text=True)
    for cell in ws[2]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        ws.column_dimensions[cell.column_letter].width = max(14, min(34, len(str(cell.value or "")) + 2))


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", type=Path, default=OUT)
    args = parser.parse_args()
    out = args.out
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "产品上架"

    columns = [
        "parent_sku",
        "handle",
        "title_en",
        "title_zh",
        "title_ja",
        "description_en_html",
        "description_zh_html",
        "description_ja_html",
        "category",
        "subcategory",
        "status",
        "tag_csv",
        "attribute_name",
    ]
    for n in range(1, 7):
        columns.extend(
            [
                f"variant{n}_sku",
                f"variant{n}_parent_sku",
                f"variant{n}_spec_value",
                f"variant{n}_price_usd",
                f"variant{n}_in_stock_0_1",
                f"variant{n}_image_url",
            ]
        )
    columns.extend(
        [
            "main_image_url",
            "gallery_image_url_1",
            "gallery_image_url_2",
            "gallery_image_url_3",
            "gallery_image_url_4",
            "gallery_image_url_5",
            "gallery_image_url_6",
            "local_image_path",
            "image_alt_en",
            "image_alt_zh",
            "image_alt_ja",
            "image_ready_0_1",
            "hide_description_0_1",
            "hide_additional_info_0_1",
            "front_hidden_0_1",
            "gallery_enabled_0_1",
            "gallery_default_index",
            "gallery_override_enabled_0_1",
            "gallery_hide_indexes",
            "slogan_en",
            "slogan_zh",
            "slogan_ja",
            "hide_slogan_0_1",
            "review_author",
            "review_email",
            "review_rating_1_5",
            "review_content",
            "review_date_yyyy_mm_dd",
            "review_status",
            "ops_note",
        ]
    )

    ws.append([""] * len(columns))
    ws.append([LABELS.get(column, column) for column in columns])
    sample = {
        "parent_sku": "TZ-QZ-002",
        "handle": "tz-qz-002",
        "title_en": "Inline Carp Fishing Weights for Carp Rig System",
        "title_zh": "欧鲤钓内联铅坠",
        "title_ja": "カープ用インラインウェイト",
        "description_en_html": "<p>Woo Description (EN) ...</p>",
        "description_zh_html": "<p>中文描述...</p>",
        "description_ja_html": "<p>日本語説明...</p>",
        "category": "铅坠 (Sinkers)",
        "status": "发布 (publish)",
        "tag_csv": "批发 wholesale",
        "attribute_name": "重量 (Weight)",
        "main_image_url": "https://carp-ybb.com/wp-content/uploads/2026/07/TZ-QZ-002.jpeg",
        "local_image_path": "D:\\product-images\\TZ-QZ-002-main.jpg",
        "image_alt_en": "Inline carp fishing weights",
        "image_alt_zh": "欧鲤钓内联铅坠",
        "image_alt_ja": "カープ用インラインウェイト",
        "image_ready_0_1": 1,
        "hide_description_0_1": 0,
        "hide_additional_info_0_1": 0,
        "front_hidden_0_1": 0,
        "gallery_enabled_0_1": 1,
        "gallery_default_index": 0,
        "gallery_override_enabled_0_1": 0,
        "hide_slogan_0_1": 0,
        "ops_note": "示例行，正式交付前替换；生成时间 " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    for index, spec_price in enumerate([("56g", 0.49), ("71g", 0.59), ("85g", 0.69), ("99g", 0.79), ("113g", 0.89)], 1):
        spec, price = spec_price
        sample[f"variant{index}_sku"] = f"TZ-QZ-002-{spec}"
        sample[f"variant{index}_parent_sku"] = "TZ-QZ-002"
        sample[f"variant{index}_spec_value"] = spec
        sample[f"variant{index}_price_usd"] = price
        sample[f"variant{index}_in_stock_0_1"] = 1
    ws.append([sample.get(column, "") for column in columns])

    group_ranges = [
        (1, 13, "Woo 基础 + 三语描述"),
        (14, 49, "变体 1-6"),
        (50, 61, "图片资产"),
        (62, 72, "YBB 站点管理覆盖"),
        (73, len(columns), "评价可选 + 备注"),
    ]
    _style_group_row(ws, group_ranges)
    _style_header_row(ws)
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 58
    for row in ws.iter_rows(min_row=3, max_row=200):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _apply_list(ws, _field_range(columns, "category"), CATEGORIES, allow_blank=False)
    _apply_list(ws, _field_range(columns, "status"), STATUS_OPTIONS, allow_blank=False)
    _apply_list(ws, _field_range(columns, "review_status"), REVIEW_STATUS_OPTIONS)
    _apply_list(ws, _field_range(columns, "review_rating_1_5"), ["1", "2", "3", "4", "5"])

    yes_no_fields = [
        *(f"variant{index}_in_stock_0_1" for index in range(1, 7)),
        "image_ready_0_1",
        "hide_description_0_1",
        "hide_additional_info_0_1",
        "front_hidden_0_1",
        "gallery_enabled_0_1",
        "gallery_override_enabled_0_1",
        "hide_slogan_0_1",
    ]
    for field in yes_no_fields:
        _apply_list(ws, _field_range(columns, field), ["0", "1"])

    wb.save(out)
    print(str(out))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
