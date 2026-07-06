#!/usr/bin/env python3
"""Import Taizhou euro-carp Excel taxonomy into catalog-taxonomy.json."""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

HANDLE_MAP = {
    "2026新品": ("2026-new-products", "2026 New Products"),
    "铅坠": ("sinkers", "Sinkers"),
    "饵笼": ("bait-cages", "Bait Cages"),
    "线组": ("rigs", "Rigs"),
    "铅坠钓组": ("sinker-rigs", "Sinker Rigs"),
    "饵笼钓组": ("bait-cage-rigs", "Bait Cage Rigs"),
    "欧鲤鱼钩": ("carp-hooks", "Carp Hooks"),
    "套盒-欧鲤�?: ("euro-carp-kits", "Euro Carp Kits"),
    "配件-金属": ("accessories-metal", "Metal Accessories"),
    "配件-塑料": ("accessories-plastic", "Plastic Accessories"),
    "支架": ("rod-pod-accessories", "Rod Pods & Stands"),
    "周边设备": ("peripheral-equipment", "Peripheral Equipment"),
}

SKIP_SHEETS = {"WpsReserved_CellImgList"}
DEFAULT_XLSX = Path.home() / "Desktop" / "2026泰州欧鲤钓类目表_白底净�?xlsx"
OUT_PATH = Path(__file__).resolve().parents[1] / "lib" / "data" / "catalog-taxonomy.json"


def parse_sheet(ws) -> list[str]:
    types: list[str] = []
    seen: set[str] = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        cell = row[0] if row else None
        if cell is None:
            continue
        name = str(cell).strip()
        if not name or name.startswith("="):
            continue
        if name not in seen:
            seen.add(name)
            types.append(name)
    return types


def main() -> int:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        print(f"Excel not found: {xlsx}", file=sys.stderr)
        return 1

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    categories = []
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        if sheet_name not in HANDLE_MAP:
            print(f"Skipping unknown sheet: {sheet_name}", file=sys.stderr)
            continue
        handle, title_en = HANDLE_MAP[sheet_name]
        product_types = parse_sheet(wb[sheet_name])
        categories.append(
            {
                "sheetName": sheet_name,
                "handle": handle,
                "titleCn": sheet_name,
                "titleEn": title_en,
                "productTypeCount": len(product_types),
                "productTypes": product_types,
            }
        )

    main_categories = categories[:8]
    other_children = categories[8:]
    taxonomy = {
        "source": str(xlsx),
        "mainCategories": main_categories,
        "other": {
            "handle": "other",
            "titleCn": "其他",
            "titleEn": "Other",
            "children": other_children,
            "productTypeCount": sum(c["productTypeCount"] for c in other_children),
        },
        "navOrder": [c["handle"] for c in main_categories]
        + ["other", "oem-odm"],
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(
        json.dumps(taxonomy, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUT_PATH}")
    print(f"Main: {len(main_categories)}, Other children: {len(other_children)}")
    for child in other_children:
        print(f"  {child['titleCn']}: {child['productTypeCount']} types")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
