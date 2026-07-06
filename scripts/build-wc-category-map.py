#!/usr/bin/env python3
"""Build WooCommerce product_cat sync payload from Taizhou Excel taxonomy."""

from __future__ import annotations

import json
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

# Reuse sheet map from import-taizhou-categories.py
HANDLE_MAP = {
    "2026新品": ("2026-new-products", "2026 New Products"),
    "铅坠": ("sinkers", "Sinkers"),
    "饵笼": ("bait-cages", "Bait Cages"),
    "线组": ("rigs", "Rigs"),
    "铅坠钓组": ("sinker-rigs", "Sinker Rigs"),
    "饵笼钓组": ("bait-cage-rigs", "Bait Cage Rigs"),
    "欧鲤鱼钩": ("carp-hooks", "Carp Hooks"),
    "套盒-欧鲤�?: ("euro-carp-kits", "Euro Carp Kits"),
    "配件-金属": ("accessories-metal", "Metal Accessories"),
    "配件-塑料": ("accessories-plastic", "Plastic Accessories"),
    "支架": ("rod-pod-accessories", "Rod Pods & Stands"),
    "周边设备": ("peripheral-equipment", "Peripheral Equipment"),
}

SKIP_SHEETS = {"WpsReserved_CellImgList"}
MAIN_SHEET_COUNT = 8
DEFAULT_XLSX = Path.home() / "Desktop" / "2026泰州欧鲤钓类目表_白底净�?xlsx"
OUT_PATH = Path(__file__).resolve().parents[1] / "lib" / "data" / "wc-category-sync.json"

LEGACY_PRODUCT_CAT_SLUGS = [
    "carp-fishing-leads",
    "carp-fishing-rigs",
    "carp-fishing-feeder",
    "carp-fishing-accessories",
    "carp-fishing-tools-add-ons",
    "carp-fishing-rod-supports",
    "carp-fishing-lines-hooks",
    "feeder-fishing-rig",
    "lead-fishing-rig",
    "inline-tube-insert-lead",
    "swivel-lead-lead-with-swivel",
    "fishing-box",
]


def slugify_type(parent_handle: str, type_name: str, index: int) -> str:
    base = re.sub(r"\s+", "-", type_name.strip().lower())
    base = re.sub(r"[^\w\-]+", "", base, flags=re.UNICODE)
    base = base.strip("-")[:40]
    if not base:
        base = f"type-{index}"
    return f"{parent_handle}--{base}"


def parse_sheet_rows(ws) -> tuple[list[dict], dict[str, list[str]]]:
    """Return product types and sku -> [parent_slug, type_slug] paths."""
    types: list[dict] = []
    sku_paths: dict[str, list[str]] = {}
    seen_types: set[str] = set()
    current_type: str | None = None
    current_type_slug: str | None = None
    type_index = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        col_a = row[0] if row and len(row) > 0 else None
        col_d = row[3] if row and len(row) > 3 else None

        if col_a is not None and str(col_a).strip() and not str(col_a).strip().startswith("="):
            current_type = str(col_a).strip()
            if current_type not in seen_types:
                seen_types.add(current_type)
                type_index += 1
                current_type_slug = None  # set by caller

        if col_d is None:
            continue
        sku_base = str(col_d).strip()
        if not sku_base or sku_base == "货号":
            continue
        if current_type and sku_base not in sku_paths:
            sku_paths[sku_base] = {"typeName": current_type}

    return types, sku_paths


def main() -> int:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        print(f"Excel not found: {xlsx}", file=sys.stderr)
        return 1

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    terms: list[dict] = []
    sku_to_slugs: dict[str, list[str]] = {}

    sheet_order = [s for s in wb.sheetnames if s not in SKIP_SHEETS and s in HANDLE_MAP]
    main_sheets = sheet_order[:MAIN_SHEET_COUNT]
    other_sheets = sheet_order[MAIN_SHEET_COUNT:]

    # Other parent
    terms.append(
        {
            "slug": "other",
            "nameEn": "Other",
            "nameZh": "其他",
            "parent": None,
        }
    )

    def add_main_sheet(sheet_name: str, parent_override: str | None = None) -> str:
        handle, title_en = HANDLE_MAP[sheet_name]
        parent_slug = parent_override
        ws = wb[sheet_name]

        if parent_override is None:
            terms.append(
                {
                    "slug": handle,
                    "nameEn": title_en,
                    "nameZh": sheet_name,
                    "parent": None,
                }
            )
            parent_for_types = handle
        else:
            terms.append(
                {
                    "slug": handle,
                    "nameEn": title_en,
                    "nameZh": sheet_name,
                    "parent": parent_override,
                }
            )
            parent_for_types = handle

        seen: set[str] = set()
        current_type: str | None = None
        current_type_slug: str | None = None
        type_index = 0

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            col_a = row[0] if row else None
            col_d = row[3] if row and len(row) > 3 else None

            if col_a is not None and str(col_a).strip() and not str(col_a).strip().startswith("="):
                current_type = str(col_a).strip()
                if current_type not in seen:
                    seen.add(current_type)
                    type_index += 1
                    current_type_slug = slugify_type(parent_for_types, current_type, type_index)
                    terms.append(
                        {
                            "slug": current_type_slug,
                            "nameEn": current_type,
                            "nameZh": current_type,
                            "parent": parent_for_types,
                        }
                    )

            if not col_d:
                continue
            sku_base = str(col_d).strip()
            if not sku_base or sku_base == "货号":
                continue
            if not current_type or not current_type_slug:
                continue

            path = (
                [parent_override, parent_for_types, current_type_slug]
                if parent_override
                else [parent_for_types, current_type_slug]
            )
            sku_to_slugs[sku_base] = path

        return handle

    for sheet_name in main_sheets:
        add_main_sheet(sheet_name)

    for sheet_name in other_sheets:
        add_main_sheet(sheet_name, parent_override="other")

    payload = {
        "source": str(xlsx),
        "taxonomy": "product_cat",
        "preserve": ["product_brand", "product_tag", "pa_*"],
        "removeLegacySlugs": LEGACY_PRODUCT_CAT_SLUGS,
        "skuPrefixFallback": {
            "TZ-XP": ["2026-new-products"],
            "TZ-HK": ["carp-hooks"],
        },
        "terms": terms,
        "skuToCategorySlugs": sku_to_slugs,
        "stats": {
            "termCount": len(terms),
            "skuBaseCount": len(sku_to_slugs),
            "mainCategories": len(main_sheets),
            "otherChildren": len(other_sheets),
        },
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_PATH}")
    print(json.dumps(payload["stats"], ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
